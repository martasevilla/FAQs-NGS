#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera faqs.html a partir de la plantilla (faqs.template.html) i del ZIP
de panells (Panells.zip).

No cal tocar aquest fitxer per actualitzar els panells: només cal
substituir Panells.zip al repositori. El GitHub Action executa aquest
script automàticament i torna a construir faqs.html.

Ús manual (si mai el vols executar tu):
    pip install openpyxl
    python generar_panells.py
"""

import os
import re
import sys
import json
import zipfile
import tempfile
import datetime
import warnings

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("Falta la llibreria 'openpyxl'. Instal·la-la amb: pip install openpyxl")

# ---------------------------------------------------------------------------
# Configuració
# ---------------------------------------------------------------------------
ZIP_PATH = "Panells.zip"
TEMPLATE_PATH = "faqs.template.html"
OUTPUT_PATH = "faqs.html"
PLACEHOLDER = "/*__PANELLS_DATA__*/"

# Nom oficial de cada grup de malalties (full "Abast" de 00.Llistat_panells.xlsx)
ABAST = {
    'NRL': 'Neuropediatria i neurodesenvolupament',
    'GAS': 'Gastroenterologia',
    'MET': 'Malalties metabòliques i mitocondrials',
    'MUS': 'Malalties neuromusculars',
    'GEN': 'Síndromes dismòrfics',
    'ONC': 'Oncologia',
    'CAR': 'Cardiologia',
    'NEF': 'Nefrologia',
    'IMM': 'Immunologia',
    'END': 'Endocrinologia',
    'OFT': 'Oftalmologia',
    'HEM': 'Hematologia',
    'ORL': 'Otorrinolaringologia',
    'DER': 'Dermatologia',
    'OBS': 'Medicina fetal',
    'CIR': 'Cirurgia',
    'REU': 'Reumatologia',
    'PNE': 'Pneumologia',
    'OSS': 'Malalties òssies',
}

# Abreviatures per a la columna "Origen"
ORIG = {
    'PanelApp_UK': 'UK',
    'PanelApp_AUS': 'AUS',
    'HPO': 'HPO',
    'Bibliografia': 'Bib',
    'Altres': 'Altres',
}


def abbr(origen):
    if not origen:
        return ''
    return ';'.join(ORIG.get(p.strip(), p.strip()) for p in str(origen).split(';'))


def fdate(d):
    if isinstance(d, (datetime.datetime, datetime.date)):
        return d.strftime('%Y-%m-%d')
    return ''


def clean_name(filename):
    """01.Discapacitat_intellectual_v2.xlsx -> ('Discapacitat intellectual', 'v2')"""
    n = re.sub(r'^\d+\.\s*', '', filename)      # treu "01." inicial
    n = re.sub(r'\.xlsx$', '', n)
    m = re.search(r'_(v\d+)$', n)
    ver = m.group(1) if m else ''
    n = re.sub(r'_v\d+$', '', n)
    n = n.replace('_', ' ').strip()
    return n, ver


def build_data(root):
    """Recorre la carpeta 'Panells' extreta i retorna l'estructura de dades."""
    folders = sorted(
        d for d in os.listdir(root)
        if os.path.isdir(os.path.join(root, d))
        and re.match(r'\d\d\.', d)
        and not d.startswith('99')          # exclou 99.OBSOLETS
    )

    groups = []
    n_panels = n_genes = 0
    for folder in folders:
        num, code = folder.split('.')[0], folder.split('.')[1]
        fpath = os.path.join(root, folder)
        panels = []
        for f in sorted(os.listdir(fpath)):
            if f.startswith('00.'):          # plantilles internes
                continue
            full = os.path.join(fpath, f)
            if f.endswith('.xlsx'):
                try:
                    wb = openpyxl.load_workbook(full, data_only=True, read_only=True)
                except Exception as e:
                    print(f"  AVÍS: no s'ha pogut llegir {f}: {e}")
                    continue
                if 'Gens' not in wb.sheetnames:
                    wb.close()
                    continue
                rows = list(wb['Gens'].iter_rows(values_only=True))
                wb.close()
                genes = [
                    [r[0], abbr(r[1]), fdate(r[3] if len(r) > 3 else None)]
                    for r in rows[1:] if r and r[0]
                ]
                name, ver = clean_name(f)
                panels.append({'name': name, 'version': ver, 'n': len(genes), 'genes': genes})
                n_panels += 1
                n_genes += len(genes)
            elif f.endswith('.url'):         # dreceres a panells d'un altre grup
                name, ver = clean_name(re.sub(r'\.xlsx$', '', f))
                panels.append({'name': name, 'version': ver, 'alias': True})
        if not panels:
            continue
        groups.append({'num': num, 'code': code, 'name': ABAST.get(code, code), 'panels': panels})

    return {'updated': datetime.date.today().strftime('%Y-%m-%d'), 'groups': groups}, n_panels, n_genes


def find_panells_root(extract_dir):
    """Localitza la carpeta que conté els grups (01.NRL, 02.GAS, ...)."""
    # cas típic: <extract_dir>/Panells/01.NRL...
    for dirpath, dirnames, _ in os.walk(extract_dir):
        if any(re.match(r'\d\d\.[A-Z]{3}$', d) for d in dirnames):
            return dirpath
    return extract_dir


def main():
    if not os.path.exists(ZIP_PATH):
        sys.exit(f"No es troba {ZIP_PATH}")
    if not os.path.exists(TEMPLATE_PATH):
        sys.exit(f"No es troba {TEMPLATE_PATH}")

    with tempfile.TemporaryDirectory() as tmp:
        with zipfile.ZipFile(ZIP_PATH) as z:
            z.extractall(tmp)
        root = find_panells_root(tmp)
        data, n_panels, n_genes = build_data(root)

    data_js = "window.PANELLS=" + json.dumps(data, ensure_ascii=False, separators=(',', ':')) + ";"

    template = open(TEMPLATE_PATH, encoding='utf-8').read()
    if PLACEHOLDER not in template:
        sys.exit(f"La plantilla no conté el marcador {PLACEHOLDER}")
    html = template.replace(PLACEHOLDER, data_js)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as fh:
        fh.write(html)

    print(f"OK · {len(data['groups'])} grups · {n_panels} panells · {n_genes} gens")
    print(f"Escrit {OUTPUT_PATH} ({len(html.encode('utf-8')):,} bytes)")


if __name__ == '__main__':
    main()
